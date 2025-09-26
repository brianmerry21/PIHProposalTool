import pandas as pd
import openpyxl
import os
import logging
import re
from openpyxl.utils import get_column_letter
from openpyxl.styles import Border, Side, Font, Alignment
from openpyxl.styles import Border, Side, Font, Alignment, PatternFill


# Set up logging
logger = logging.getLogger(__name__)



def apply_borders_to_sheet(sheet):
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )

    for row in sheet.iter_rows():
        for cell in row:
            cell.border = thin_border



def analyze_excel_template(template_path):
    """
    Analyze an Excel template file to understand its structure
    
    Args:
        template_path: Path to the Excel template file
        
    Returns:
        A dictionary containing information about the template structure
    """
    try:
        # Load the workbook to get sheet names and structure info
        wb = openpyxl.load_workbook(template_path)
        sheet_names = wb.sheetnames
        
        template_info = {
            "file_path": template_path,
            "sheet_names": sheet_names,
            "sheets_info": {}
        }
        
        # Analyze each sheet
        for sheet_name in sheet_names:
            sheet = wb[sheet_name]
            
            # Get sheet dimensions
            max_row = sheet.max_row
            max_col = sheet.max_column
            
            # Collect information about all cells for better analysis
            headers = {}
            content_cells = {}
            
            # Scan all rows and columns to find headers and content
            for row in range(1, min(max_row, 20) + 1):  # Limit to first 20 rows for efficiency
                for col in range(1, max_col + 1):
                    cell_value = sheet.cell(row=row, column=col).value
                    if cell_value:
                        if row == 1:
                            headers[get_column_letter(col)] = cell_value
                        else:
                            # Store non-empty cell values for content analysis
                            content_cells[f"{get_column_letter(col)}{row}"] = cell_value
            
            # Look for key cells that might contain formulas or special formatting
            key_cells = {}
            for row in range(1, max_row + 1):
                for col in range(1, max_col + 1):
                    cell = sheet.cell(row=row, column=col)
                    # Check for formulas
                    # Check for formulas
                    if isinstance(cell.value, str) and cell.value.startswith('='):
                        key_cells[cell.coordinate] = {
                            "type": "formula",
                            "formula": cell.value,
                            "value": None  # or keep evaluated value if you want
                        }

                    # Check for merged cells
                    for merged_range in sheet.merged_cells.ranges:
                        if cell.coordinate in merged_range:
                            key_cells[cell.coordinate] = {
                                "type": "merged",
                                "range": str(merged_range),
                                "value": cell.value
                            }
                            break
            
            template_info["sheets_info"][sheet_name] = {
                "dimensions": {"rows": max_row, "columns": max_col},
                "headers": headers,
                "key_cells": key_cells,
                "content_cells": content_cells
            }
        
        return template_info
    
    except Exception as e:
        logger.error(f"Error analyzing Excel template: {str(e)}")
        raise

def create_excel_from_template(template_path, output_path, data_mapping):
    """
    Create a new Excel file based on a template and populate it with data
    
    Args:
        template_path: Path to the Excel template file
        output_path: Path to save the new Excel file
        data_mapping: Dictionary mapping cell references to data values
    """
    try:
        # Load the template workbook
        wb = openpyxl.load_workbook(template_path)
        
        # Apply data mapping to populate cells and insert new rows as needed
        for sheet_name, cell_data in data_mapping.items():
            if sheet_name in wb.sheetnames:
                sheet = wb[sheet_name]
                
                # Check if we have item rows to insert
                if "__item_rows" in cell_data:
                    item_rows = cell_data["__item_rows"]
                    if item_rows:
                        logger.info(f"Inserting {len(item_rows)} new rows for line items")
                        
                        # Determine where to insert rows based on the "Machine" section header
                        # The template suggests starting insert after row 10
                        insert_row_index = 11
                        
                        # Insert rows and populate them
                        for i, item in enumerate(item_rows):
                            # Insert new row
                            sheet.insert_rows(insert_row_index + i)
                            
                            # Populate the cells with the extracted data
                            sheet.cell(row=insert_row_index + i, column=2).value = item.get('Description', '')
                            sheet.cell(row=insert_row_index + i, column=3).value = item.get('Total List', '')

                            # Handle the INCLUDED/NOT INCLUDED flags
                            status = item.get('Total List', '').upper().strip()
                            if 'INCLUDED' in status:
                                sheet.cell(row=insert_row_index + i, column=4).value = 'X'
                                sheet.cell(row=insert_row_index + i, column=5).value = 'X'
                            elif 'NOT INCLUDED' in status:
                                sheet.cell(row=insert_row_index + i, column=6).value = 'X'
                            
                            # Apply proper formatting to the new row
                            try:
                                thin_border = Border(
                                    left=Side(style='thin'), right=Side(style='thin'), 
                                    top=Side(style='thin'), bottom=Side(style='thin')
                                )
                                
                                for col in range(1, sheet.max_column + 1):
                                    target_cell = sheet.cell(row=insert_row_index + i, column=col)
                                    target_cell.border = thin_border
                                    if col == 2: # Apply alignment for Description column
                                        target_cell.alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
                            except Exception as style_e:
                                logger.warning(f"Could not apply cell formatting: {style_e}")
                
                # Apply regular cell mapping
                for cell_ref, value in cell_data.items():
                    if cell_ref.startswith("__"):
                        continue

                    if isinstance(value, str) and value.startswith('='):
                        # It's a formula
                        sheet[cell_ref].value = value
                    else:
                        # Regular value
                        sheet[cell_ref].value = value
            else:
                logger.warning(f"Sheet '{sheet_name}' not found in template")
        
        
        # ... after filling in all cell values and formatting ...

        # Apply borders to all sheets
        for sheet_name in wb.sheetnames:
            sheet = wb[sheet_name]
            apply_borders_to_sheet(sheet)

        # Save the workbook
        wb.save(output_path)
        logger.debug(f"Excel file created at {output_path}")

        return output_path
    
    except Exception as e:
        logger.error(f"Error creating Excel from template: {str(e)}")
        raise

# def extract_mapping_from_pdf(pdf_text, template_info, mapping_rules):
#     """
#     Extract data from PDF text and create a mapping for Excel template
#     """
#     logger = logging.getLogger(__name__)
#     data_mapping = {}

#     # -----------------------------
#     # Extract specific items from pages 16-17
#     # -----------------------------
#     # try:
#     #     specific_text = ""
#     #     if pdf_text and len(pdf_text) >= 17:
#     #         specific_text = "\n".join(pdf_text[15:17])
#     #     elif pdf_text and len(pdf_text) >= 16:
#     #         specific_text = "\n".join(pdf_text[15:])
#     #     else:
#     #         specific_text = "\n".join(pdf_text) if pdf_text else ""

#     #     # Isolate only the block we care about
#     #     block_match = re.search(
#     #         r"(MODULA®LIFT VLM Units Model:.*?Standard Unit Warranty - 2 Years - Parts & Labor.*)",
#     #         specific_text,
#     #         re.DOTALL,
#     #     )
#     #     if block_match:
#     #         specific_text = block_match.group(1)

#     #     sheet_names_list = list(mapping_rules.keys())
#     #     if sheet_names_list:
#     #         main_sheet = sheet_names_list[0]
#     #         item_list = []
#     #         current_category = "Machine"

#     #         # Regex: description + qty + status/price
#     #         item_pattern = re.compile(
#     #             r"^(.*?)(?:\s+(\d+(?:\.\d+)?))?\s+(INCLUDED|NOT INCLUDED|By End User|N/A|[$]\s*[\d,\.]+)$",
#     #             re.IGNORECASE | re.MULTILINE,
#     #         )

#     #         for line in specific_text.splitlines():
#     #             line = line.strip()
#     #             if not line:
#     #                 continue  # skip blanks

#     #             match = item_pattern.match(line)
#     #             if match:
#     #                 description, qty, status_price = match.groups()
#     #                 description = description.strip()
#     #                 qty = qty.strip() if qty else ""
#     #                 status_price = status_price.strip()

#     #                 # Skip if it's just "INCLUDED" with no description
#     #                 if description.upper() == "" and status_price.upper() == "INCLUDED":
#     #                     continue

#     #                 # Default qty = 1 if missing
#     #                 if not qty and status_price.upper() == "INCLUDED":
#     #                     qty = "1"

#     #                 item_data = {
#     #                     "Description": description,
#     #                     "Qty": qty,
#     #                     "Total List": status_price.replace(" ", ""),
#     #                     "Category": current_category,
#     #                 }
#     #                 item_list.append(item_data)

#     #         data_mapping[main_sheet] = {"__item_rows": item_list}
#     #         logger.debug(f"Extracted {len(item_list)} clean items for {main_sheet}")

#     # except Exception as e:
#     #     logger.error(f"Error extracting specific items: {str(e)}")
        
        
#     # try:
#     #     sheet_names_list = list(mapping_rules.keys())
#     #     if sheet_names_list:
#     #         main_sheet = sheet_names_list[0]

#     #         # -----------------------------
#     #         # Hardcoded items for pages 16-17
#     #         # -----------------------------
#     #         # hardcoded_items = [
#     #         #     {"Description": "Single Internal Bay Work Station", "Qty": "1", "Total List": "123", "Category": "Machine"},
#     #         #     {"Description": "Large Unit Base Plates", "Qty": "1", "Total List": "789", "Category": "Machine"}
#     #         # ]

#     #         data_mapping[main_sheet] = {"__item_rows": hardcoded_items}
#     #         logger.debug(f"Added {len(hardcoded_items)} hardcoded items for {main_sheet}")

#     # except Exception as e:
#     #     logger.error(f"Error adding hardcoded items for pages 16-17: {str(e)}")

        
#     # -----------------------------
#     # Extract Optional Items from page 18 (REVISED BLOCK)
#     # -----------------------------
#     try:
#         if pdf_text and len(pdf_text) >= 18:
#             optional_text = pdf_text[17]  # Page 18 text
#             sheet_name = list(mapping_rules.keys())[0]

#             # Predefined mapping of optional items to Excel rows
#             predefined_rows = {
#                 "Sliding Operator Console": 63,
#                 "Wireless Bar Code Reader (1D)": 65,
#                 "Wireless Bar Code Reader (2D)": 67,
#                 "RFID Badge Reader": 69,
#                 "Magnetic Badge Reader": 71,
#                 "ESD Protection System": 73,
#                 "Spare Parts Kit - Level A": 75,
#                 "Spare Parts Kit - Level B": 77,
#                 "Spare Parts Kit - Level C": 79,
#                 "Premium 24/5 TSA (Annual Fee)": 81,
#                 "Premium 24/7 TSA (Annual Fee)": 83,
#                 "Tray Partitions and Dividers": 95,
#                 "Put-to-light System": 96,
#             }

#             # Hardcoded values for some items
#             # hardcoded_values = {
#             #     "Sliding Operator Console": "1535.00",  # Hardcoded
#             #     "Wireless Bar Code Reader (1D)": "1986.00",  
#             #     "ESD Protection System": "1329.00",
#             #     "Premium 24/5 TSA (Annual Fee)": "5000.00",
#             #     "Premium 24/7 TSA (Annual Fee)": "8250.00"
#             # }

#             # Regex: capture item name and number
#             number_pattern = re.compile(r"^(.*?)\s+([\d,]+\.\d{2}|TBD)$", re.MULTILINE)

#             for predefined_name, row_number in predefined_rows.items():
#                 cell_ref = f"E{row_number}"
#                 if sheet_name not in data_mapping:
#                     data_mapping[sheet_name] = {}

#                 # If a hardcoded value exists, use it
#                 if predefined_name in hardcoded_values:
#                     data_mapping[sheet_name][cell_ref] = hardcoded_values[predefined_name]
#                     logger.debug(f"Hardcoded {predefined_name} -> {cell_ref} = {hardcoded_values[predefined_name]}")
#                     continue  # Skip extraction from PDF

#                 # Otherwise, try extracting from page 18
#                 for match in number_pattern.finditer(optional_text):
#                     item_name = match.group(1).strip()
#                     item_value = match.group(2).strip()

#                     if item_value.upper() == "TBD":
#                         continue

#                     item_value_clean = item_value.replace(",", "")
#                     if predefined_name.lower() in item_name.lower():
#                         data_mapping[sheet_name][cell_ref] = item_value_clean
#                         logger.debug(f"Mapped {item_name} -> {cell_ref} = {item_value_clean}")

#     except Exception as e:
#         logger.error(f"Error extracting optional items from page 18: {str(e)}")




#     # -----------------------------
#     # Extract remaining fields from mapping_rules
#     # -----------------------------
#     full_text = "\n".join(pdf_text)

#     for sheet_name, cell_rules in mapping_rules.items():
#         if sheet_name not in data_mapping:
#             data_mapping[sheet_name] = {}

#         for cell_ref, rule in cell_rules.items():
#             if cell_ref == "__item_rows__":
#                 continue

#             value = ""
#             if "pattern" not in rule and "custom_handler" not in rule and "default" in rule:
#                 value = rule["default"]
#             elif "custom_handler" in rule:
#                 try:
#                     value = rule["custom_handler"](full_text)
#                 except Exception as e:
#                     logger.error(f"Custom handler failed for {cell_ref}: {e}")
#                     value = rule.get("default", "")
#             elif "pattern" in rule:
#                 match = re.search(rule["pattern"], full_text, re.MULTILINE)
#                 value = match.group(rule.get("group", 0)) if match else rule.get("default", "")
#                 if "transform" in rule and callable(rule["transform"]):
#                     try:
#                         value = rule["transform"](value)
#                     except Exception as e:
#                         logger.error(f"Transform failed for {cell_ref}: {e}")
#                         pass

#             data_mapping[sheet_name][cell_ref] = value

#     return data_mapping

def extract_mapping_from_pdf(pdf_text, template_info, mapping_rules):
    """
    Extract data from PDF text and create a mapping for Excel template
    """
    logger = logging.getLogger(__name__)
    data_mapping = {}

    sheet_names_list = list(mapping_rules.keys())
    if sheet_names_list:
        main_sheet = sheet_names_list[0]

        # -----------------------------
        # Hardcoded items go into existing columns (do NOT insert rows)
        # -----------------------------
        # Map Description -> Column B, Qty -> Column C, Total List -> Column E, etc.
        hardcoded_items = [
            # {"Description": "Single Internal Bay Work Station", "Category": "Machine"},
            # {"Description": "Large Unit Base Plates", "Category": "Machine"},
        ]

        # Map each item to a specific row based on template structure
        # Example: rows 11, 12... or any rows the template already reserves
        start_row = 11
        for idx, item in enumerate(hardcoded_items):
            row = start_row + idx
            if main_sheet not in data_mapping:
                data_mapping[main_sheet] = {}

            # Fill columns without inserting rows
            data_mapping[main_sheet][f"B{row}"] = item.get("Description", "")
            data_mapping[main_sheet][f"C{row}"] = item.get("Qty", "")
            data_mapping[main_sheet][f"E{row}"] = item.get("Total List", "")

            # Optional: handle INCLUDED/NOT INCLUDED flags in specific columns
            status = item.get("Total List", "").upper()
            if "INCLUDED" in status:
                data_mapping[main_sheet][f"D{row}"] = "X"
                data_mapping[main_sheet][f"F{row}"] = "X"
            elif "NOT INCLUDED" in status:
                data_mapping[main_sheet][f"G{row}"] = "X"

        logger.debug(f"Mapped {len(hardcoded_items)} hardcoded items into template rows without shifting formulas.")

    # hardcoded software items
    # Hardcoded software items
    hardcoded_software_items = [
        {"Description": "Modula WMS Base Software", "Category": "Software"},
        {"Description": "SW On-Site Installation and Training", "Category": "Software"},
        {"Description": "PIH On-Site Software Support", "Category": "Software"}
    ]

    software_start_row = 18  # Row where software items start

    for idx, item in enumerate(hardcoded_software_items):
        # Skip 3rd item to go in the 4th row
        row = software_start_row + idx
        if idx == 2:
            row += 1

        if main_sheet not in data_mapping:
            data_mapping[main_sheet] = {}

        # Description column (B)
        data_mapping[main_sheet][f"B{row}"] = item.get("Description", "")

        # # Number column (K) for last two items
        # if idx > 0:
        #     data_mapping[main_sheet][f"K{row}"] = 1

        # Extra value in a far-away column 
        # if idx ==1:
        #     data_mapping[main_sheet][f"G{row}"] = "4054.00"

        # Optional flags or other columns
        data_mapping[main_sheet][f"C{row}"] = item.get("Qty", "")
        data_mapping[main_sheet][f"F{row}"] = item.get("Total List", "")


        # Optional flags (if needed)
        status = item.get("Total List", "").upper()
        if "INCLUDED" in status:
            data_mapping[main_sheet][f"D{row}"] = "X"
            data_mapping[main_sheet][f"G{row}"] = "X"
        elif "NOT INCLUDED" in status:
            data_mapping[main_sheet][f"H{row}"] = "X"


    # -----------------------------
    # Extract Optional Items from page 18 (existing logic)
    # -----------------------------
    # try:
    #     if pdf_text and len(pdf_text) >= 18:
    #         optional_text = pdf_text[17]  # Page 18 text
    #         sheet_name = main_sheet

    #         predefined_rows = {
    #             "Sliding Operator Console": 63,
    #             "Wireless Bar Code Reader (1D)": 65,
    #             "Wireless Bar Code Reader (2D)": 67,
    #             "RFID Badge Reader": 69,
    #             "Magnetic Badge Reader": 71,
    #             "ESD Protection System": 73,
    #             "Spare Parts Kit - Level A": 75,
    #             "Spare Parts Kit - Level B": 77,
    #             "Spare Parts Kit - Level C": 79,
    #             "Premium 24/5 TSA (Annual Fee)": 81,
    #             "Premium 24/7 TSA (Annual Fee)": 83,
    #             "Tray Partitions and Dividers": 95,
    #             "Put-to-light System": 96,
    #         }

    #         # hardcoded_values = {
    #         #     "Sliding Operator Console": "1535.00",
    #         #     "Wireless Bar Code Reader (1D)": "1986.00",
    #         #     "ESD Protection System": "1329.00",
    #         #     "Premium 24/5 TSA (Annual Fee)": "5000.00",
    #         #     "Premium 24/7 TSA (Annual Fee)": "8250.00"
    #         # }

    #         number_pattern = re.compile(r"^(.*?)\s+([\d,]+\.\d{2}|TBD)$", re.MULTILINE)

    #         for name, row_number in predefined_rows.items():
    #             cell_ref = f"E{row_number}"
    #             if sheet_name not in data_mapping:
    #                 data_mapping[sheet_name] = {}

    #             # Use hardcoded value if exists
    #             if name in hardcoded_values:
    #                 data_mapping[sheet_name][cell_ref] = hardcoded_values[name]
    #                 continue

    #             # Otherwise, extract from PDF
    #             for match in number_pattern.finditer(optional_text):
    #                 item_name = match.group(1).strip()
    #                 item_value = match.group(2).strip()
    #                 if item_value.upper() == "TBD":
    #                     continue
    #                 if name.lower() in item_name.lower():
    #                     data_mapping[sheet_name][cell_ref] = item_value.replace(",", "")
    # except Exception as e:
    #     logger.error(f"Error extracting optional items from page 18: {str(e)}")

    # -----------------------------
    # Extract remaining fields from mapping_rules
    # -----------------------------
    full_text = "\n".join(pdf_text)
    for sheet_name, cell_rules in mapping_rules.items():
        if sheet_name not in data_mapping:
            data_mapping[sheet_name] = {}

        for cell_ref, rule in cell_rules.items():
            if cell_ref.startswith("__"):
                continue

            value = ""
            if "pattern" not in rule and "custom_handler" not in rule and "default" in rule:
                value = rule["default"]
            elif "custom_handler" in rule:
                try:
                    value = rule["custom_handler"](full_text)
                except Exception as e:
                    logger.error(f"Custom handler failed for {cell_ref}: {e}")
                    value = rule.get("default", "")
            elif "pattern" in rule:
                match = re.search(rule["pattern"], full_text, re.MULTILINE)
                value = match.group(rule.get("group", 0)) if match else rule.get("default", "")
                if "transform" in rule and callable(rule["transform"]):
                    try:
                        value = rule["transform"](value)
                    except Exception as e:
                        logger.error(f"Transform failed for {cell_ref}: {e}")
                        pass

            data_mapping[sheet_name][cell_ref] = value

    return data_mapping



def process_pdf_to_excel_with_template(pdf_path, template_path, excel_path, mapping_rules=None, customer_info=None):
    """
    Process a PDF file and use its data to populate an Excel template
    
    Args:
        pdf_path: Path to the PDF file
        template_path: Path to the Excel template file
        excel_path: Path to save the Excel file
        mapping_rules: Optional dictionary defining rules to map PDF data to Excel cells
            If None, default rules will be created based on template analysis
        customer_info: Optional dictionary with customer information to override PDF extraction
    """
    try:
        from utils.pdf_processor import extract_text_from_pdf
        
        # Extract text from the PDF
        pdf_text = extract_text_from_pdf(pdf_path)
        
        # Analyze the Excel template
        template_info = analyze_excel_template(template_path)
        
        # If no mapping rules provided, create default rules
        if mapping_rules is None:
            mapping_rules = create_default_mapping_rules(template_info)
        
        # Extract data from PDF text and create mapping for Excel template
        data_mapping = extract_mapping_from_pdf(pdf_text, template_info, mapping_rules)
        
        # Create Excel file from template with extracted data
        create_excel_from_template(template_path, excel_path, data_mapping)
        
        logger.debug(f"Excel file created from template at {excel_path}")
        
    except Exception as e:
        logger.error(f"Error processing PDF to Excel with template: {str(e)}")
        raise

def create_default_mapping_rules(template_info):
    """
    Create default mapping rules based on template analysis
    
    Args:
        template_info: Dictionary containing template structure information
    
    Returns:
        Dictionary defining rules to map PDF data to Excel cells
    """
    mapping_rules = {}
    
    for sheet_name, sheet_info in template_info["sheets_info"].items():
        mapping_rules[sheet_name] = {}
        
        for col, header in sheet_info["headers"].items():
            cell_ref = f"{col}2"
            pattern = f"{header}:\\s*([\\w\\d\\s.]+)"
            
            mapping_rules[sheet_name][cell_ref] = {
                "pattern": pattern,
                "group": 1
            }
    
    return mapping_rules

def analyze_template_and_save(template_path, output_path):
    """
    Analyze an Excel template and save the analysis to a text file
    
    Args:
        template_path: Path to the Excel template file
        output_path: Path to save the analysis text file
    """
    try:
        # Analyze the template
        template_info = analyze_excel_template(template_path)
        
        # Format the analysis as text
        analysis_text = []
        analysis_text.append(f"Excel Template Analysis: {os.path.basename(template_path)}")
        analysis_text.append("=" * 80)
        analysis_text.append(f"Number of sheets: {len(template_info['sheet_names'])}")
        analysis_text.append(f"Sheet names: {', '.join(template_info['sheet_names'])}")
        analysis_text.append("")
        
        for sheet_name, sheet_info in template_info["sheets_info"].items():
            analysis_text.append(f"Sheet: {sheet_name}")
            analysis_text.append("-" * 40)
            analysis_text.append(f"Dimensions: {sheet_info['dimensions']['rows']} rows x {sheet_info['dimensions']['columns']} columns")
            
            if sheet_info["headers"]:
                analysis_text.append("Headers:")
                for col, header in sheet_info["headers"].items():
                    analysis_text.append(f"  Column {col}: {header}")
            
            if sheet_info["key_cells"]:
                analysis_text.append("Key Cells:")
                for cell_ref, cell_info in sheet_info["key_cells"].items():
                    analysis_text.append(f"  Cell {cell_ref}: Type: {cell_info['type']}, Value: {cell_info['value']}")
                    
                    if cell_info["type"] == "formula":
                        analysis_text.append(f"    Formula: {cell_info['formula']}")
                    elif cell_info["type"] == "merged":
                        analysis_text.append(f"    Merged Range: {cell_info['range']}")
            
            if "content_cells" in sheet_info and sheet_info["content_cells"]:
                analysis_text.append("Content Cells (first 20 rows):")
                for cell_ref, value in sheet_info["content_cells"].items():
                    if len(str(value)) > 50:
                        # Truncate long values for clarity
                        display_value = str(value)[:47] + "..."
                    else:
                        display_value = value
                    analysis_text.append(f"  Cell {cell_ref}: {display_value}")
            
            analysis_text.append("")
        
        # Save the analysis to a text file
        with open(output_path, 'w') as f:
            f.write('\n'.join(analysis_text))
        
        logger.debug(f"Template analysis saved to {output_path}")
        
    except Exception as e:
        logger.error(f"Error analyzing template and saving: {str(e)}")
        raise